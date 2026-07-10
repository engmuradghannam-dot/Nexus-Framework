"""Import the controls library from the master Excel workbooks.

Idempotent: safe to run repeatedly (uses update_or_create on natural keys).

Usage:
    python manage.py import_controls
    python manage.py import_controls --data-dir /path/to/xlsx
"""

import os

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.controls.models import (
    AIAgent,
    FormControl,
    Industry,
    IndustryControl,
    MasterEntity,
)

# Map the workbook's status labels (Arabic + emoji) to model enum values.
STATUS_MAP = {
    "✅ موجود": FormControl.STATUS_PRESENT,
    "❌ ناقص": FormControl.STATUS_MISSING,
    "🔄 مخطط": FormControl.STATUS_PLANNED,
}


def _s(value):
    """Normalise a cell to a trimmed string."""
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Import the controls library from the master Excel workbooks."

    def add_arguments(self, parser):
        parser.add_argument(
            "--data-dir",
            default=str(settings.BASE_DIR / "data" / "controls"),
            help="Directory containing the master .xlsx workbooks.",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.stderr.write("openpyxl is required. Run: pip install openpyxl")
            return

        data_dir = options["data_dir"]
        industry_file = os.path.join(data_dir, "Industry_Control_Library_Master.xlsx")
        controls_file = os.path.join(data_dir, "AllControls_Master.xlsx")

        counts = {}
        with transaction.atomic():
            counts["industries"] = self._import_industries(industry_file)
            counts["industry_controls"] = self._import_industry_controls(industry_file)
            counts["ai_agents"] = self._import_ai_agents(industry_file)
            counts["master_entities"] = self._import_master_entities(industry_file)
            counts["form_controls"] = self._import_form_controls(controls_file)

        self.stdout.write(
            self.style.SUCCESS(
                "✅ Controls imported: "
                + ", ".join(f"{k}={v}" for k, v in counts.items())
            )
        )

    # ── loaders ──────────────────────────────────────────────

    def _sheet_rows(self, path, sheet):
        import openpyxl

        if not os.path.exists(path):
            self.stderr.write(f"⚠️  File not found, skipping: {path}")
            return []
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            self.stderr.write(f"⚠️  Sheet '{sheet}' not in {os.path.basename(path)}")
            wb.close()
            return []
        rows = list(wb[sheet].iter_rows(values_only=True))
        wb.close()
        return rows[1:]  # skip header

    def _import_industries(self, path):
        n = 0
        for row in self._sheet_rows(path, "Industry Catalog"):
            code = _s(row[0])
            if not code:
                continue
            Industry.objects.update_or_create(
                code=code,
                defaults={
                    "name": _s(row[1]),
                    "category": _s(row[2]),
                    "description": _s(row[3]),
                },
            )
            n += 1
        return n

    def _import_industry_controls(self, path):
        n = 0
        for row in self._sheet_rows(path, "Industry Control Library"):
            control_id = _s(row[2])
            if not control_id:
                continue
            IndustryControl.objects.update_or_create(
                control_id=control_id,
                defaults={
                    "industry": _s(row[0]),
                    "module": _s(row[1]),
                    "control_name": _s(row[3]),
                    "sub_control": _s(row[4]),
                    "description": _s(row[5]),
                    "required": _s(row[6]).lower() in ("yes", "true", "1", "نعم"),
                    "ai_agent": _s(row[7]),
                    "database_entity": _s(row[8]),
                    "kpi": _s(row[9]),
                    "compliance": _s(row[10]),
                },
            )
            n += 1
        return n

    def _import_ai_agents(self, path):
        n = 0
        for row in self._sheet_rows(path, "AI Agent Registry"):
            name = _s(row[1])
            if not name:
                continue
            AIAgent.objects.update_or_create(
                name=name,
                industry=_s(row[0]),
                defaults={
                    "responsibility": _s(row[2]),
                    "database_entity": _s(row[3]),
                },
            )
            n += 1
        return n

    def _import_master_entities(self, path):
        n = 0
        for row in self._sheet_rows(path, "Entity Master Library"):
            name = _s(row[0])
            if not name:
                continue
            MasterEntity.objects.update_or_create(
                name=name,
                defaults={"entity_type": _s(row[1]), "usage": _s(row[2])},
            )
            n += 1
        return n

    def _import_form_controls(self, path):
        n = 0
        for row in self._sheet_rows(path, "All Controls Master"):
            form_name = _s(row[1])
            input_name = _s(row[2])
            if not form_name or not input_name:
                continue
            try:
                seq = int(row[0]) if row[0] is not None else 0
            except (ValueError, TypeError):
                seq = 0
            FormControl.objects.update_or_create(
                form_name=form_name,
                input_name=input_name,
                defaults={
                    "seq": seq,
                    "input_type": _s(row[3]),
                    "status": STATUS_MAP.get(_s(row[4]), FormControl.STATUS_MISSING),
                    "priority": _s(row[5]) or "Medium",
                },
            )
            n += 1
        return n
